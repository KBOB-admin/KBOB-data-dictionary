import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / 'scripts' / 'validator'))

from validate_strukturvorlage import Validator  # noqa: E402


class DataTemplateCardinalityTests(unittest.TestCase):
    SOURCE = ROOT / 'templates' / 'test_files' / 'Strukturvorlage_DataDictionary_KBOB_FM_v0.9.5.xlsx'

    def test_repeated_id_represents_one_template_with_multiple_groups(self):
        workbook = load_workbook(self.SOURCE, data_only=True, read_only=True)
        sheet = workbook['Data_Template']
        rows = [
            (sheet.cell(row, 2).value, sheet.cell(row, 3).value)
            for row in range(5, sheet.max_row + 1)
            if sheet.cell(row, 2).value == 'gebaeude-dt'
        ]
        workbook.close()

        self.assertGreaterEqual(len(rows), 2)
        self.assertGreaterEqual(len({group for _identifier, group in rows if group}), 2)

    def test_validator_rejects_id_reused_for_different_classes(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / self.SOURCE.name
            shutil.copy2(self.SOURCE, path)
            workbook = load_workbook(path)
            sheet = workbook['Data_Template']
            sheet.cell(7, 1).value = 'Gebäude'
            sheet.cell(7, 2).value = 'grundstueck-dt'
            workbook.save(path)
            workbook.close()

            validator = Validator(path)
            validator.validate_matrix()
            codes = {finding.code for finding in validator.findings}
            self.assertIn('matrix_data_template_class_mismatch', codes)


if __name__ == '__main__':
    unittest.main()
