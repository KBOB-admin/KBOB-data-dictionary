import sys
import unittest
import tempfile
from pathlib import Path
from openpyxl import Workbook

# Ensure repo root is on sys.path so the validator package can be imported
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from P_workspace_codat3_validator.scripts.validator.validate_strukturvorlage import Validator


class TestDataTemplateParser(unittest.TestCase):
    def test_parse_real_areamgmt_template(self):
        # locate the AreaMgmt test template in a few likely locations
        fname = 'Strukturvorlage_AreaMgmt_v1.0.0_new.xlsx'
        candidates = [
            Path(__file__).resolve().parents[2] / 'templates' / 'test_files' / fname,
            Path(__file__).resolve().parents[3] / 'P_workspace_codat3_validator' / 'templates' / 'test_files' / fname,
            Path('/home/Dave/.openclaw/workspace-datadict/P_workspace_codat3_validator/templates/test_files') / fname,
        ]
        p = None
        for c in candidates:
            if c.exists():
                p = c.resolve()
                break
        self.assertIsNotNone(p, f'Required test file missing; checked: {candidates}')
        v = Validator(p)
        structure = v.parse_datatemplate_structure()
        # Known layout values for the AreaMgmt example
        self.assertEqual(structure['property_anchor'], 5)
        self.assertEqual(structure['related_doc_anchor'], 55)
        self.assertEqual(structure['related_doc_item_anchor'], 61)
        self.assertEqual(structure['property_start_col'], 6)
        self.assertEqual(structure['property_end_col'], 54)
        self.assertGreaterEqual(len(structure['doc_relation_pairs']), 1)

    def test_parse_shifted_anchors(self):
        # Build a small workbook with shifted anchors to test dynamic detection
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_path = Path(tmp.name)
        tmp.close()
        wb = Workbook()
        # remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet('Data_Template')
        # Write row 2 with anchors in shifted positions
        ws.cell(row=2, column=3).value = 'Property - Designation/Bezeichnung/Désignation/Designazione'
        # property columns 4..8
        for c in range(4, 9):
            ws.cell(row=2, column=c).value = f'Prop{c}'
        ws.cell(row=2, column=9).value = 'RelatedDocumentName (EN)'
        # filler columns
        ws.cell(row=2, column=11).value = 'Filler'
        ws.cell(row=2, column=12).value = 'RelatedDocumentItemReference'
        ws.cell(row=2, column=13).value = 'DocSection1'
        ws.cell(row=2, column=14).value = 'DocSection2'
        wb.save(tmp_path)

        v = Validator(tmp_path)
        structure = v.parse_datatemplate_structure()
        self.assertEqual(structure['property_anchor'], 3)
        self.assertEqual(structure['related_doc_anchor'], 9)
        self.assertEqual(structure['related_doc_item_anchor'], 12)
        self.assertEqual(structure['property_start_col'], 4)
        self.assertEqual(structure['property_end_col'], 8)
        self.assertGreaterEqual(len(structure['doc_relation_pairs']), 1)

        tmp_path.unlink()


if __name__ == '__main__':
    unittest.main()
