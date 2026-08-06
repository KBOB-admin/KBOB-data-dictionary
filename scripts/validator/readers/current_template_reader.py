from __future__ import annotations

from dataclasses import replace
from pathlib import Path
from urllib.parse import quote
import re
import openpyxl

from .excel_reader import (
    DataDictionary,
    DictionaryMeta,
    DDAllowedValue,
    DDClass,
    DDClassProperty,
    DDConceptRelation,
    DDProperty,
)


def slugify(value: str) -> str:
    s = (value or '').strip().lower()
    repl = {
        'ä': 'ae', 'ö': 'oe', 'ü': 'ue', 'ß': 'ss',
        'à': 'a', 'á': 'a', 'â': 'a', 'ã': 'a', 'å': 'a',
        'è': 'e', 'é': 'e', 'ê': 'e', 'ë': 'e',
        'ì': 'i', 'í': 'i', 'î': 'i', 'ï': 'i',
        'ò': 'o', 'ó': 'o', 'ô': 'o', 'õ': 'o',
        'ù': 'u', 'ú': 'u', 'û': 'u',
        'ç': 'c', 'ñ': 'n',
        '_': '-',
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r'[^a-z0-9-]+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _split_values(raw: str | None) -> list[str]:
    if not raw:
        return []
    txt = str(raw).strip()
    if not txt:
        return []
    parts = [p.strip().strip('"') for p in re.split(r'[;,]', txt)]
    return [p for p in parts if p]


def _safe_uri(base: str, code: str) -> str:
    return f"{base}{quote(code, safe='')}"


def _normalize_ifc_entity(entity: str | None, predefined: str | None) -> str | None:
    entity = _str(entity)
    predefined = _str(predefined)
    if not entity:
        return None
    if predefined and entity.endswith(predefined):
        candidate = entity[:-len(predefined)]
        if candidate.startswith('Ifc') and len(candidate) > 3:
            return candidate
    return entity


def _row_has_meaningful_content(row) -> bool:
    meaningful = []
    for v in row:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        meaningful.append(s)
    if not meaningful:
        return False
    joined = ' '.join(meaningful).lower()
    guidance_markers = [
        'required human-readable', 'required canonical', 'system-generated', 'validator should',
        'should reference', 'should match', 'dropdown liste', 'validierung', 'referenziert / system generiert',
        'erforderlich für excel', 'sprachen'
    ]
    return not any(marker in joined for marker in guidance_markers)


def _iter_row_dicts(ws, data_start: int) -> list[dict[str, object]]:
    headers = [_str(c.value) or '' for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if _row_has_meaningful_content(row):
            rows.append(dict(zip(headers, row)))
    return rows


def _default_meta(path: Path) -> DictionaryMeta:
    stem = path.stem.lower()
    if 'kbob' in stem:
        code = 'KBOB'
    elif 'bdch' in stem:
        code = 'BdCH'
    else:
        code = 'HESEM'
    return DictionaryMeta(
        org_code=code,
        org_name_de=code,
        org_name_en=code,
        dd_uri='',
        dd_version='0.1',
        dd_status='Preview',
        raw={},
    )


def _parse_header(ws, meta: DictionaryMeta) -> DictionaryMeta:
    for row in ws.iter_rows(min_row=2, values_only=True):
        field_name = _str(row[0]) if len(row) > 0 else None
        value = _str(row[1]) if len(row) > 1 else None
        if not field_name or value is None:
            continue
        meta.raw[field_name] = value
        fl = field_name.lower()
        if fl == 'organizationcode':
            meta.org_code = value
        elif fl == 'dictionaryuri':
            meta.dd_uri = value
        elif fl in ('dictionaryversion', 'version'):
            meta.dd_version = value
        elif fl == 'lifestyclestatus' or fl == 'lifecyclestatus':
            meta.dd_status = value
    return meta


def _lindas_base(meta: DictionaryMeta) -> str:
    org1 = (meta.raw.get('OrganizationCodeLindas') or meta.raw.get('OrganizationCode') or meta.org_code or 'org').lower()
    org2 = (meta.raw.get('OrganizationSubCode') or meta.raw.get('OrganizationCode') or meta.org_code or org1).lower()
    dd_code = (meta.raw.get('DictionaryCode') or meta.raw.get('DictionaryName (EN)') or meta.raw.get('DictionaryName (DE)') or 'dd').strip()
    dd_code = slugify(dd_code).lower().replace('-', '_')
    return f'https://lindas.admin.ch/{org1}/{org2}/{dd_code}/'


def _parse_classes(ws, meta: DictionaryMeta) -> list[DDClass]:
    rows = _iter_row_dicts(ws, 7)
    base = f'{_lindas_base(meta)}class/'
    classes = []
    for r in rows:
        label_en = _str(r.get('Designation (EN)'))
        label_de = _str(r.get('Bezeichnung (DE)'))
        code = _str(r.get('Class-ID')) or slugify(label_en or label_de or '')
        if not any([code, label_en, label_de]):
            continue
        predefined = _str(r.get('PredefinedType'))
        entity_raw = _str(r.get('IfcObject Entity'))
        cls = DDClass(
            code=code,
            class_type='Class',
            name_de=label_de or label_en or '',
            name_fr=_str(r.get('Désignation (FR)')) or '',
            name_en=label_en or label_de or '',
            definition_de=_str(r.get('Beschreibung (DE)')) or '',
            definition_fr=_str(r.get('Description (FR)')) or '',
            owned_uri=_str(r.get('GUID/URI')) or _safe_uri(base, code),
            parent_class_code=_str(r.get('Class-Assignment')),
            ifc_entity_code=_normalize_ifc_entity(entity_raw, predefined),
            ifc_predefined_type=predefined,
            ifc_uri=_str(r.get('IFC_URI')),
            rds_reference=None,
            crb_code=_str(r.get('Class-Code')),
            status=_str(r.get('Status')) or 'Preview',
            version_date=_str(r.get('Version date')),
            document_reference=_str(r.get('Provenance (PROV)')),
            countries_of_use='CH',
        )
        setattr(cls, 'name_it', _str(r.get('Designazione (IT)')) or '')
        setattr(cls, 'definition_it', _str(r.get('Descrizione (IT)')) or '')
        setattr(cls, 'ifc_type_object_entity_code', _str(r.get('IfcTypeObject Entity')))
        setattr(cls, 'object_type', _str(r.get('ObjectType')))
        setattr(cls, 'related_document', _str(r.get('RelatedDocumentName (EN)')))
        classes.append(cls)
    return classes


def _parse_values(ws, meta: DictionaryMeta) -> list[dict]:
    return _iter_row_dicts(ws, 7)


def _parse_properties(ws, value_ws, meta: DictionaryMeta) -> tuple[list[DDProperty], list[DDAllowedValue]]:
    rows = _iter_row_dicts(ws, 7)
    value_rows = _parse_values(value_ws, meta) if value_ws is not None else []
    value_map = {}
    for vr in value_rows:
        raw_id = _str(vr.get('Enumeration-ID'))
        if raw_id:
            value_map[raw_id] = vr
    value_base = f'{_lindas_base(meta)}allowed-value/'
    prop_base = f'{_lindas_base(meta)}property/'
    props = []
    avs = []
    for r in rows:
        property_id = _str(r.get('Property-ID'))
        property_code = _str(r.get('Property-Code'))
        label_en = _str(r.get('Designation (EN)'))
        code = property_id or property_code or slugify(label_en or '')
        if not code:
            continue
        value_list_id = _str(r.get('Enumeration-ID')) or _str(r.get('Values.Enumeration-ID'))
        value_row = value_map.get(value_list_id or '')
        enum_values_raw = _str(value_row.get('Enumeration (EN)')) if value_row else None
        vals = _split_values(enum_values_raw)
        prop = DDProperty(
            code=code,
            name_de=_str(r.get('Bezeichnung (DE)')) or label_en or '',
            name_fr=_str(r.get('Désignation (FR)')) or '',
            name_en=label_en or _str(r.get('Bezeichnung (DE)')) or '',
            definition_de=_str(r.get('Beschreibung (DE)')) or '',
            definition_fr=_str(r.get('Description (FR)')) or '',
            owned_uri=_str(r.get('GUID/URI')) or _safe_uri(prop_base, code),
            data_type=_str(r.get('DataType\n(Base Type)')) or 'STRING',
            data_type_ifc=_str(r.get('DataType\n(IFC)')),
            property_value_kind='List' if vals else 'Single',
            unit_label=_str(r.get('Unit Name (EN)')),
            unit_qudt_iri=_str(r.get('QUDT URI')),
            enumeration_values=','.join(vals) if vals else None,
            ifc_property_uri=_str(r.get('IFC_URI')),
            ifc_pset_uri=_str(r.get('IfcPropertySet (Pset)\nIfcQuantitySet (Qto)')),
            property_set_name=None,
            rds_reference=None,
            status=_str(r.get('Status')) or 'Preview',
            version_date=_str(r.get('Version date')),
            prov_attributed_to=_str(r.get('Provenance (PROV)')),
        )
        setattr(prop, 'name_it', _str(r.get('Designazione (IT)')) or '')
        setattr(prop, 'definition_it', _str(r.get('Descrizione (IT)')) or '')
        props.append(prop)
        if vals:
            for idx, v in enumerate(vals, start=1):
                val_code = slugify(v)
                avs.append(DDAllowedValue(
                    property_code=code,
                    code=val_code,
                    value_de=v,
                    value_fr='',
                    value_en=v,
                    definition_de=None,
                    owned_uri=_safe_uri(value_base, f'{code}/{val_code}'),
                    sort_number=idx,
                    status='Preview',
                ))
    return props, avs


def _parse_groups(ws) -> list[dict]:
    return _iter_row_dicts(ws, 7)


def _parse_documents(ws, meta: DictionaryMeta) -> list[dict]:
    rows = _iter_row_dicts(ws, 7)
    docs = []
    base_doc = f'{_lindas_base(meta)}document/'
    for r in rows:
        document_label = _str(r.get('DocumentName (EN)')) or ''
        document_id = _str(r.get('Document-ID')) or slugify(document_label)
        item = dict(r)
        item['SourceCode'] = _str(r.get('DocumentOwner')) or ''
        item['Dokument-ID'] = document_id
        item['DocumentCode'] = _str(r.get('Document-Code')) or ''
        item['DocumentLabel'] = document_label
        item['DocumentName'] = document_label
        item['DocumentName (EN)'] = document_label
        item['Owner'] = _str(r.get('DocumentOwner')) or ''
        item['VersionDate'] = _str(r.get('Version date')) or ''
        item['DocumentGroupCode'] = _str(r.get('Security level/Sicherheitsstufe/Niveau de sécurité/Livello di sicurezza')) or ''
        item['DocumentGroupName'] = _str(r.get('Accessibility/Zugänglichkeit/Accessibilité/Accessibilità')) or ''
        item['Sicherheitsstufe'] = item['DocumentGroupCode']
        item['Zugänglichkeit'] = item['DocumentGroupName']
        item['OwnedUri'] = _safe_uri(base_doc, document_id)
        docs.append(item)
    return docs


def _parse_matrix(ws, properties: list[DDProperty], group_refs: set[str]) -> list[DDClassProperty]:
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 5:
        return []
    row2 = [(_str(v) or '') for v in rows[1]]
    property_anchor = None
    document_anchor = None
    governance_anchor = None
    for idx, value in enumerate(row2, start=1):
        if value == 'Property - Designation/Bezeichnung/Désignation/Designazione':
            property_anchor = idx
        elif value == 'Document - Designation/Bezeichnung/Désignation/Designazione':
            document_anchor = idx
        elif value == 'Governance':
            governance_anchor = idx
    if property_anchor is None:
        return []
    property_start_col = property_anchor + 1
    property_end_col = (document_anchor - 1) if document_anchor and document_anchor > property_start_col else (governance_anchor - 1 if governance_anchor and governance_anchor > property_start_col else len(row2))
    prop_lookup = {}
    for p in properties:
        for key in [p.code, p.name_en, p.name_de, getattr(p, 'name_it', None), p.name_fr]:
            if key:
                prop_lookup[str(key).strip()] = p
    property_cols = []
    for col_idx in range(property_start_col, property_end_col + 1):
        label = row2[col_idx - 1] if col_idx - 1 < len(row2) else ''
        prop = prop_lookup.get(label)
        if prop:
            property_cols.append((col_idx, prop.code, label))
    cps = []
    for row_idx in range(5, ws.max_row + 1):
        vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        if not _row_has_meaningful_content(vals):
            continue
        class_code = _str(ws.cell(row_idx, 1).value)
        group_label = _str(ws.cell(row_idx, 3).value)
        for col_idx, property_code, label in property_cols:
            cell = _str(ws.cell(row_idx, col_idx).value)
            if not cell:
                continue
            marker = cell.strip()
            lower = marker.lower()
            override = None if lower == 'x' else marker
            cps.append(DDClassProperty(
                class_code=class_code or '',
                property_code=property_code,
                property_set_name=group_label if group_label in group_refs else None,
                is_required=True,
                is_writable=True,
                allowed_values_override=override,
            ))
    return cps


def _parse_group_refs(ws) -> set[str]:
    refs = set()
    for r in _iter_row_dicts(ws, 7):
        for key in ['GoP-ID', 'GoP-Code', 'Designation (EN)', 'Bezeichnung (DE)', 'Désignation (FR)', 'Designazione (IT)']:
            val = _str(r.get(key))
            if val:
                refs.add(val)
    return refs


def _parse_concept_relations(ws) -> list[DDConceptRelation]:
    rows = _iter_row_dicts(ws, 7)
    crs = []
    for r in rows:
        subject_code = _str(r.get('ConceptCode')) or ''
        concept_type = _str(r.get('ConceptType'))
        related_uri = _str(r.get('RelatedConceptUri')) or ''
        relation_type = _str(r.get('RelationType')) or ''
        notes = _str(r.get('Notes'))
        if subject_code and relation_type and related_uri:
            crs.append(DDConceptRelation(
                subject_code=subject_code,
                concept_type=concept_type,
                relation_type=relation_type,
                related_uri=related_uri,
                notes=notes,
            ))
    return crs


def load_current_template_dd(path: Path) -> DataDictionary:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    meta = _parse_header(wb['Header'], _default_meta(path)) if 'Header' in wb.sheetnames else _default_meta(path)
    classes = _parse_classes(wb['Classes'], meta)
    properties, allowed_values = _parse_properties(wb['Properties'], wb['Values'], meta)
    group_sheet = 'GroupOfProperties' if 'GroupOfProperties' in wb.sheetnames else None
    group_refs = _parse_group_refs(wb[group_sheet]) if group_sheet else set()
    class_properties = _parse_matrix(wb['Data_Template'], properties, group_refs) if 'Data_Template' in wb.sheetnames else []
    concept_sheet = 'ConceptRelation' if 'ConceptRelation' in wb.sheetnames else None
    dd = DataDictionary(
        source_file=path,
        meta=meta,
        classes=classes,
        properties=properties,
        class_properties=class_properties,
        allowed_values=allowed_values,
        concept_relations=_parse_concept_relations(wb[concept_sheet]) if concept_sheet else [],
    )
    setattr(dd, 'documents', _parse_documents(wb['Documents'], meta) if 'Documents' in wb.sheetnames else [])
    setattr(dd, 'merkmalsgruppen', _parse_groups(wb[group_sheet]) if group_sheet else [])
    return dd
