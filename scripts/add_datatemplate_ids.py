#!/usr/bin/env python3
"""Add stable DataTemplate-ID support without shifting existing workbook columns.

Column B in Data_Template is an existing spacer column. This migration gives it
an explicit identity role so several rows can contribute 0..n GroupOfProperties
to one DataTemplate by repeating the same DataTemplate-ID. The conservative
``row`` strategy preserves legacy row-level template identity when aggregation
has not been semantically approved.
"""

from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def slug(value: object) -> str:
    text = str(value or '').strip().lower()
    for source, target in {'ä': 'ae', 'ö': 'oe', 'ü': 'ue', 'ß': 'ss'}.items():
        text = text.replace(source, target)
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', text)).strip('-')


def resolved_cell_value(workbook, cell) -> object:
    value = cell.value
    if not isinstance(value, str) or not value.startswith('='):
        return value
    match = re.fullmatch(r"='?([^']+)'?!([A-Z]+[0-9]+)", value.strip())
    if not match:
        return None
    sheet_name, coordinate = match.groups()
    if sheet_name not in workbook.sheetnames:
        return None
    return workbook[sheet_name][coordinate].value


def validation_signature(workbook) -> tuple:
    signatures = []
    for sheet in workbook.worksheets:
        for validation in sheet.data_validations.dataValidation:
            signatures.append((
                sheet.title,
                str(validation.sqref),
                validation.type,
                validation.formula1,
                validation.formula2,
            ))
    return tuple(sorted(signatures))


def defined_name_signature(workbook) -> tuple:
    return tuple(sorted(
        (name, item.attr_text, item.localSheetId)
        for name, item in workbook.defined_names.items()
    ))


def migrated_identifier(row: int, class_label: object, group_label: object, strategy: str) -> str:
    if strategy == 'row':
        return f'{row:04d}-{slug(class_label)}-{slug(group_label) or "no-group"}'
    return f'{slug(class_label)}-dt'


def migrate(path: Path, strategy: str = 'class', replace_existing: bool = False) -> dict:
    workbook = load_workbook(path)
    if 'Data_Template' not in workbook.sheetnames:
        workbook.close()
        return {'path': str(path), 'changed': False, 'rows': 0, 'reason': 'Data_Template missing'}
    before_validations = validation_signature(workbook)
    before_names = defined_name_signature(workbook)
    sheet = workbook['Data_Template']
    existing = [
        column for column in range(1, sheet.max_column + 1)
        if str(sheet.cell(2, column).value or '').strip() == 'DataTemplate-ID'
    ]
    if existing and existing[0] != 2:
        workbook.close()
        raise ValueError(f'{path}: DataTemplate-ID already exists outside reserved column B')

    sheet.cell(1, 2).value = 'Data Templates'
    sheet.cell(2, 2).value = 'DataTemplate-ID'
    sheet.cell(3, 2).value = 'Stable identity'
    sheet.cell(4, 2).value = 'Repeat the same ID across rows to attach 0..n GroupOfProperties'
    for row, style_source in ((1, 3), (2, 3), (3, 1), (4, 1)):
        source = sheet.cell(row, style_source)
        target = sheet.cell(row, 2)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
            target.alignment = copy(source.alignment)
    sheet.column_dimensions['B'].width = 28

    populated = 0
    for row in range(5, sheet.max_row + 1):
        class_label = resolved_cell_value(workbook, sheet.cell(row, 1))
        group_label = resolved_cell_value(workbook, sheet.cell(row, 3))
        row_has_content = any(
            resolved_cell_value(workbook, sheet.cell(row, column)) not in (None, '')
            for column in range(1, sheet.max_column + 1)
            if column != 2
        )
        if not row_has_content or (class_label in (None, '') and group_label in (None, '')):
            continue
        current = str(sheet.cell(row, 2).value or '').strip()
        if current and not replace_existing:
            if not re.fullmatch(r'[a-z0-9]+(?:-[a-z0-9]+)*', current):
                workbook.close()
                raise ValueError(f'{path}: invalid existing DataTemplate-ID {current!r} in row {row}')
            populated += 1
            continue
        if class_label in (None, ''):
            workbook.close()
            raise ValueError(f'{path}: cannot derive DataTemplate-ID for populated row {row} without Class')
        identifier = migrated_identifier(row, class_label, group_label, strategy)
        if not re.fullmatch(r'[a-z0-9]+(?:-[a-z0-9]+)*', identifier):
            workbook.close()
            raise ValueError(f'{path}: cannot derive DataTemplate-ID for row {row}')
        sheet.cell(row, 2).value = identifier
        populated += 1

    workbook.save(path)
    workbook.close()

    verified = load_workbook(path)
    after_validations = validation_signature(verified)
    after_names = defined_name_signature(verified)
    if after_validations != before_validations:
        verified.close()
        raise RuntimeError(f'{path}: Excel data-validation bindings changed during migration')
    if after_names != before_names:
        verified.close()
        raise RuntimeError(f'{path}: workbook defined names changed during migration')
    if verified['Data_Template'].cell(2, 2).value != 'DataTemplate-ID':
        verified.close()
        raise RuntimeError(f'{path}: DataTemplate-ID header did not persist')
    verified.close()
    return {'path': str(path), 'changed': True, 'rows': populated}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--strategy',
        choices=('class', 'row'),
        default='class',
        help='class aggregates rows by Class; row preserves one legacy DataTemplate per populated row',
    )
    parser.add_argument(
        '--replace-existing',
        action='store_true',
        help='replace existing DataTemplate-ID values using the selected strategy',
    )
    parser.add_argument('workbooks', nargs='+', type=Path)
    args = parser.parse_args()
    for workbook in args.workbooks:
        if not workbook.is_file():
            raise FileNotFoundError(workbook)
        print(migrate(workbook, strategy=args.strategy, replace_existing=args.replace_existing))


if __name__ == '__main__':
    main()
