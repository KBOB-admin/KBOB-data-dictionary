#!/usr/bin/env python3
"""
Script to upgrade Strukturvorlage templates from v0.9.5 to v1.0.0
Adds new document relationship columns:
- RelatedDocumentName
- RelatedDocumentItemReference

Applies to sheets:
- Classes
- Properties
- Values
- GroupOfProperties
- Data_Template
"""

import argparse
import shutil
from pathlib import Path
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")


def copy_sheet_styles(source_sheet, target_sheet):
    """Copy styles from source sheet to target sheet"""
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)


def get_column_index(ws, header_row, column_name):
    """Find column index by header name"""
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value and str(cell_value).strip() == column_name:
            return col
    return None


def add_document_columns(ws, sheet_name, header_row=1, guidance_row=6):
    """Add RelatedDocumentName and RelatedDocumentItemReference columns"""
    
    # Find the rightmost column with content
    max_col = ws.max_column
    
    # Try to find existing RelatedDocumentName (EN) column
    existing_doc_name_col = None
    for col in range(1, max_col + 1):
        header = ws.cell(header_row, col).value
        if header and 'RelatedDocumentName' in str(header):
            existing_doc_name_col = col
            break
    
    # If RelatedDocumentName already exists, just add RelatedDocumentItemReference after it
    if existing_doc_name_col:
        insert_col = existing_doc_name_col + 1
        # Check if RelatedDocumentItemReference already exists
        next_header = ws.cell(header_row, insert_col).value
        if next_header and 'RelatedDocumentItemReference' in str(next_header):
            print(f"  Sheet {sheet_name}: RelatedDocument columns already exist, skipping")
            return
    else:
        # Find where to insert (after Provenance or Status or at the end)
        insert_col = max_col + 1
        for col in range(1, max_col + 1):
            header = ws.cell(header_row, col).value
            if header:
                header_str = str(header).strip()
                # Insert after Provenance (PROV) column
                if 'Provenance' in header_str or 'PROV' in header_str:
                    insert_col = col + 1
    
    # Insert RelatedDocumentName (EN) column
    ws.insert_cols(insert_col)
    ws.cell(header_row, insert_col).value = "RelatedDocumentName (EN)"
    ws.cell(header_row, insert_col).font = Font(bold=True)
    
    # Add guidance in row 6
    if sheet_name == "Classes":
        ws.cell(guidance_row, insert_col).value = "Reference to a document from the Documents sheet"
    elif sheet_name == "Properties":
        ws.cell(guidance_row, insert_col).value = "Reference to a document from the Documents sheet"
    elif sheet_name == "Values":
        ws.cell(guidance_row, insert_col).value = "Reference to a document from the Documents sheet"
    elif sheet_name == "GroupOfProperties":
        ws.cell(guidance_row, insert_col).value = "Reference to a document from the Documents sheet"
    elif sheet_name == "Data_Template":
        ws.cell(guidance_row, insert_col).value = "Reference to a document from the Documents sheet"
    
    # Insert RelatedDocumentItemReference column after RelatedDocumentName
    insert_col2 = insert_col + 1
    ws.insert_cols(insert_col2)
    ws.cell(header_row, insert_col2).value = "RelatedDocumentItemReference"
    ws.cell(header_row, insert_col2).font = Font(bold=True)
    
    # Add guidance
    ws.cell(guidance_row, insert_col2).value = "Specific section or clause within the related document"
    
    print(f"  Sheet {sheet_name}: Added RelatedDocument columns at positions {insert_col}, {insert_col2}")


def copy_impressum_sheet(source_wb, target_wb):
    """Copy Impressum sheet from source workbook to target workbook"""
    if 'Impressum' not in source_wb.sheetnames:
        print("  Warning: No Impressum sheet found in source")
        return False
    
    source_impressum = source_wb['Impressum']
    
    # Remove existing Impressum if present
    if 'Impressum' in target_wb.sheetnames:
        del target_wb['Impressum']
    
    # Create new Impressum sheet
    target_impressum = target_wb.create_sheet('Impressum', 0)
    
    # Copy cell values
    for row in source_impressum.iter_rows():
        for cell in row:
            target_cell = target_impressum[cell.coordinate]
            target_cell.value = cell.value
    
    # Copy styles
    copy_sheet_styles(source_impressum, target_impressum)
    
    # Copy column dimensions
    for col in range(1, source_impressum.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in source_impressum.column_dimensions:
            target_impressum.column_dimensions[col_letter].width = \
                source_impressum.column_dimensions[col_letter].width
    
    # Copy row dimensions
    for row in range(1, source_impressum.max_row + 1):
        if row in source_impressum.row_dimensions:
            target_impressum.row_dimensions[row].height = \
                source_impressum.row_dimensions[row].height
    
    # Copy merged cells
    for merged_range in source_impressum.merged_cells.ranges:
        target_impressum.merge_cells(str(merged_range))
    
    print("  Copied Impressum sheet")
    return True


def add_document_item_reference_column(wb):
    """Add DocumentItemReference column to Documents sheet"""
    if 'Documents' not in wb.sheetnames:
        print("  Warning: No Documents sheet found")
        return False
    
    ws = wb['Documents']
    header_row = 1
    guidance_row = 6
    
    # Find the maximum column with content
    max_col = ws.max_column
    
    # Find Provenance (PROV) column
    provenance_col = None
    for col in range(1, max_col + 1):
        header = ws.cell(header_row, col).value
        if header and 'Provenance' in str(header):
            provenance_col = col
            break
    
    if not provenance_col:
        print("  Warning: Could not find Provenance column in Documents")
        return False
    
    # Insert DocumentItemReference column after Provenance
    insert_col = provenance_col + 1
    ws.insert_cols(insert_col)
    
    # Set header
    ws.cell(header_row, insert_col).value = "DocumentItemReference"
    ws.cell(header_row, insert_col).font = Font(bold=True)
    
    # Add guidance row
    ws.cell(guidance_row, insert_col).value = 'Specific section or clause reference (e.g., "3.2.1", "Clause 4.5")'
    
    print(f"  Sheet Documents: Added DocumentItemReference column at position {insert_col}")
    return True


def update_header_sheet(ws):
    """Update the Header sheet with new version information"""
    # Find and update version cell
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row, 1).value
        if cell_value and str(cell_value).strip().lower() == "version":
            # Update version to 1.0.0
            version_cell = ws.cell(row, 2)
            if version_cell.value and "v0.9.5" in str(version_cell.value):
                version_cell.value = str(version_cell.value).replace("v0.9.5", "v1.0.0")
            elif version_cell.value and "0.9.5" in str(version_cell.value):
                version_cell.value = str(version_cell.value).replace("0.9.5", "1.0.0")
            else:
                version_cell.value = "1.0.0"
            print(f"  Updated version to 1.0.0")
            break


def upgrade_template(template_path, output_path, impressum_source_path=None):
    """Upgrade a single template file"""
    print(f"\nProcessing: {template_path}")
    
    # Load workbook
    wb = openpyxl.load_workbook(template_path)
    
    # Update Header sheet version
    if 'Header' in wb.sheetnames:
        update_header_sheet(wb['Header'])
    
    # Add DocumentItemReference to Documents sheet
    add_document_item_reference_column(wb)
    
    # Add document columns to reference sheets
    sheets_to_update = ['Classes', 'Properties', 'Values', 'GroupOfProperties', 'Data_Template']
    
    for sheet_name in sheets_to_update:
        if sheet_name in wb.sheetnames:
            # Determine header row based on sheet
            if sheet_name == 'Data_Template':
                header_row = 2  # Data_Template has headers in row 2
                guidance_row = 4
            else:
                header_row = 1
                guidance_row = 6
            
            add_document_columns(wb[sheet_name], sheet_name, header_row, guidance_row)
        else:
            print(f"  Warning: Sheet {sheet_name} not found")
    
    # Copy Impressum if source provided
    if impressum_source_path and impressum_source_path.exists():
        source_wb = openpyxl.load_workbook(impressum_source_path)
        copy_impressum_sheet(source_wb, wb)
        source_wb.close()
    
    # Save upgraded workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    
    print(f"  Saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Upgrade Strukturvorlage templates from v0.9.5 to v1.0.0"
    )
    parser.add_argument(
        "--templates-dir",
        type=Path,
        default=Path(__file__).parent.parent / "templates",
        help="Directory containing template files"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Output directory for upgraded templates (defaults to same as templates-dir)"
    )
    parser.add_argument(
        "--impressum-source",
        type=Path,
        help="Path to template file containing the Impressum to copy"
    )
    
    args = parser.parse_args()
    
    templates_dir = args.templates_dir
    output_dir = args.output_dir or templates_dir
    impressum_source = args.impressum_source
    
    print(f"Templates directory: {templates_dir}")
    print(f"Output directory: {output_dir}")
    if impressum_source:
        print(f"Impressum source: {impressum_source}")
    
    # Process empty template
    empty_template = templates_dir / "Strukturvorlage_DataDictionary_empty_v0.9.5.xlsx"
    if empty_template.exists():
        output_path = output_dir / "Strukturvorlage_DataDictionary_empty_v1.0.0.xlsx"
        upgrade_template(empty_template, output_path, impressum_source)
    else:
        print(f"Warning: {empty_template} not found")
    
    # Process public empty template
    public_empty_template = templates_dir / "Strukturvorlage_DataDictionary_empty_public_v0.9.5.xlsx"
    if public_empty_template.exists():
        output_path = output_dir / "Strukturvorlage_DataDictionary_empty_public_v1.0.0.xlsx"
        upgrade_template(public_empty_template, output_path, impressum_source)
    else:
        print(f"Warning: {public_empty_template} not found")
    
    # Process AreaMgmt template
    test_files_dir = templates_dir / "test_files"
    area_mgmt_template = test_files_dir / "Strukturvorlage_AreaMgmt_v0.5.0.xlsx"
    if area_mgmt_template.exists():
        output_path = test_files_dir / "Strukturvorlage_AreaMgmt_v1.0.0.xlsx"
        # Use the AreaMgmt v1.0.0 impressum as source for itself (already updated)
        impressum_for_area = test_files_dir / "Strukturvorlage_AreaMgmt_v1.0.0.xlsx"
        impressum_src = impressum_for_area if impressum_for_area.exists() else None
        upgrade_template(area_mgmt_template, output_path, impressum_src)
    else:
        print(f"Warning: {area_mgmt_template} not found")
    
    # Process KBOB template
    kbob_template = test_files_dir / "Strukturvorlage_DataDictionary_KBOB_FM_v1.0.0.xlsx"
    if kbob_template.exists():
        output_path = test_files_dir / "Strukturvorlage_DataDictionary_KBOB_FM_v1.0.1.xlsx"
        upgrade_template(kbob_template, output_path, impressum_source)
    else:
        print(f"Warning: {kbob_template} not found")
    
    print("\n✓ Template upgrade complete!")


if __name__ == "__main__":
    main()
